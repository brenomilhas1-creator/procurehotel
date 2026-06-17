#!/usr/bin/env python3
"""
Importa tabelas reais do user (Alpha Food, Gergran, Makro, Lusigel)
para a base de dados.

Cada tabela tem produtos com preços REAIS do user.
Nenhum dado é inventado — tudo vem das células dos ficheiros.
"""
import openpyxl
import re
import os
import json
import sys
from pathlib import Path

# Adicionar o caminho para chamar psql depois
UPLOADS = Path('/tmp/uploads')

def extract_name(s):
    """Limpar nome de produto"""
    if not s:
        return None
    s = str(s).strip()
    # Remover trailing '-' e espaços
    s = re.sub(r'\s*-\s*$', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s[:255] if s else None

def parse_number(v):
    """Parse number from various formats"""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    # Remove currency symbols and spaces
    s = re.sub(r'[^\d,.\-]', '', s)
    # PT format: 1.234,56 (thousands ., decimal ,)
    # US format: 1,234.56 (thousands ,, decimal .)
    # Tentar US primeiro (se tem ambos)
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            # PT: 1.234,56
            s = s.replace('.', '').replace(',', '.')
        else:
            # US: 1,234.56
            s = s.replace(',', '')
    elif ',' in s:
        # Pode ser decimal PT ou thousands US
        parts = s.split(',')
        if len(parts) == 2 and len(parts[1]) <= 2:
            s = s.replace(',', '.')
        else:
            s = s.replace(',', '')
    try:
        return float(s)
    except:
        return None

def detect_unit(s):
    """Detectar unidade do texto"""
    if not s:
        return 'un'
    s_lower = str(s).lower()
    if any(x in s_lower for x in [' kg', 'kg ', 'quilo', '(kg)']):
        return 'kg'
    if any(x in s_lower for x in [' lt', ' l ', 'litro', '(l)', 'ml']):
        return 'l'
    if any(x in s_lower for x in [' cx', 'caixa', '(cx)']):
        return 'cx'
    if any(x in s_lower for x in [' un', 'unid']):
        return 'un'
    return 'un'

def find_unit_in_text(s):
    """Procurar unidade em texto descritivo (ex: '5 KG', 'CX 12', '1L')"""
    if not s:
        return None
    s = str(s).upper()
    # Procurar padrões
    m = re.search(r'(\d+)\s*KG', s)
    if m:
        return ('kg', float(m.group(1)))
    m = re.search(r'(\d+)\s*L[TG]?\b', s)
    if m:
        return ('l', float(m.group(1)))
    m = re.search(r'CX\s*(\d+)', s)
    if m:
        return ('cx', float(m.group(1)))
    return None

# ============================
# TABELA ALPHA FOOD
# Formato: no, estab, ref, design, net_atual, PREÇO CONC, NOME
# ============================
def process_alpha_food():
    """Alpha Food - padrão do cliente. Cria produtos a partir de design + preços."""
    wb = openpyxl.load_workbook(UPLOADS / '1781264386228-tabela_ALPHA_FOOD.xlsx', data_only=True)
    ws = wb.active
    products = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 3:
            continue
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        # Coluna 0: no, 1: estab, 2: ref, 3: design, 4: net_atual (preço), 5: PREÇO CONC, 6: NOME
        ref = str(row[2] or '').strip() if row[2] else None
        design = extract_name(row[3])
        net_atual = parse_number(row[4])  # preço cliente
        preco_conc = parse_number(row[5])  # preço concorrente (outros fornecedores)
        nome = str(row[6] or '').strip() if row[6] else None
        if not design:
            continue
        # Criar product só se temos design
        products.append({
            'master_name': design[:255],
            'sku': ref[:64] if ref else None,
            'price_alpha_food': net_atual,  # preço de Alpha Food
            'price_concorrente': preco_conc,
            'concorrente': nome,
            'source': 'ALPHA FOOD',
            'source_ref': 'tabela ALPHA FOOD.xlsx',
        })
    return products

# ============================
# TABELA GERGRAN
# Formato: REF, DESCRIÇÃO, UND. CAIXA, PESO, Preço un 2026, Preço cx 2026, IVA
# ============================
def process_gergran():
    """Gergran - pães e pastelaria. Tabela completa com preços 2026."""
    wb = openpyxl.load_workbook(UPLOADS / '1781264379613-Tabela_Gergran.xlsx', data_only=True)
    ws = wb.active
    products = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 5:
            continue
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        # Verificar se é header de categoria (sem ref mas tem texto)
        ref = str(row[1] or '').strip() if row[1] else None
        desc = extract_name(row[2])
        und_caixa = parse_number(row[3])
        peso = parse_number(row[4])
        preco_un = parse_number(row[5])  # preço unitário
        preco_cx = parse_number(row[6])  # preço caixa
        iva = parse_number(row[7])
        # Filtrar headers (sem ref ou ref não numérico)
        if not ref or not ref.isdigit():
            continue
        if not desc or not preco_un:
            continue
        products.append({
            'master_name': desc[:255],
            'sku': ref,
            'unit': 'un',  # Gergran vende por unidade (peso por un)
            'package_qty': und_caixa or 1,
            'price': preco_un,
            'price_cx': preco_cx,
            'tax_rate': int((iva or 0) * 100),
            'source': 'GERGRAN',
            'source_ref': 'Tabela Gergran.xlsx',
            'category': 'Padaria' if not desc.lower().startswith('mini') and ('pão' in desc.lower() or 'pão ' in desc.lower() or 'baguete' in desc.lower() or 'bolinha' in desc.lower()) else 'Pastelaria',
        })
    return products

# ============================
# TABELA MAKRO (Comparativo de Mercado)
# Formato: CÓDIGO MAKRO, DESCRIÇÃO, CATEGORIA, PREÇO ATUAL, NOVO PREÇO, Outro Fornecedor
# ============================
def process_tabela_makro():
    """Comparativo de preços Makro vs outros fornecedores."""
    wb = openpyxl.load_workbook(UPLOADS / '1781264370648-tabela_makro.xlsx', data_only=True)
    ws = wb.active
    products = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 2:
            continue
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        codigo = str(row[0] or '').strip() if row[0] else None
        descricao = extract_name(row[1])
        categoria = str(row[2] or '').strip() if row[2] else None
        preco_atual = parse_number(row[3])  # preço atual
        novo_preco = parse_number(row[4])  # novo preço
        outro_forn = str(row[5] or '').strip() if row[5] else None
        if not descricao:
            continue
        # Usar novo_preço se existir, senão preco_atual
        price = novo_preco if novo_preco is not None and novo_preco > 0 else preco_atual
        if price is None or price <= 0:
            continue
        products.append({
            'master_name': descricao[:255],
            'sku': codigo[:64] if codigo else None,
            'category': categoria,
            'price': price,
            'price_atual': preco_atual,
            'novo_preco': novo_preco,
            'outro_fornecedor': outro_forn,
            'source': 'COMPARATIVO',
            'source_ref': 'tabela makro.xlsx',
        })
    return products

# ============================
# TABELA LUSIGEL
# Formato: código, descrição, ?, ?, peso, und, preço cx, preço un
# ============================
def process_lusigel():
    """Lusigel - mini-pastelaria."""
    wb = openpyxl.load_workbook(UPLOADS / '1781264397475-lusigel.xlsx', data_only=True)
    ws = wb.active
    products = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 2:
            continue
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        codigo = str(row[0] or '').strip() if row[0] else None
        descricao = extract_name(row[1])
        peso = parse_number(row[2])
        und_cx = parse_number(row[3])  # unidades por caixa
        preco_cx = parse_number(row[5])
        preco_un = parse_number(row[6])
        if not descricao:
            continue
        if not preco_un and not preco_cx:
            continue
        price = preco_un if preco_un else (preco_cx / und_cx if und_cx else preco_cx)
        if not price or price <= 0:
            continue
        products.append({
            'master_name': descricao[:255],
            'sku': codigo[:64] if codigo else None,
            'unit': 'un',
            'package_qty': und_cx or 1,
            'price': price,
            'tax_rate': 6,
            'source': 'LUSIGEL',
            'source_ref': 'lusigel.xlsx',
            'category': 'Pastelaria',
        })
    return products

# ============================
# MAIN
# ============================
if __name__ == '__main__':
    print('=' * 60)
    print('Processando tabelas REAIS do user')
    print('=' * 60)
    
    all_data = {}
    print('\n[1/4] Alpha Food...')
    af = process_alpha_food()
    print(f'  → {len(af)} items')
    all_data['alpha_food'] = af[:5]  # sample
    
    print('\n[2/4] Gergran...')
    gg = process_gergran()
    print(f'  → {len(gg)} items')
    all_data['gergran'] = gg[:5]
    
    print('\n[3/4] Makro (Comparativo)...')
    mk = process_tabela_makro()
    print(f'  → {len(mk)} items')
    all_data['makro'] = mk[:5]
    
    print('\n[4/4] Lusigel...')
    lg = process_lusigel()
    print(f'  → {len(lg)} items')
    all_data['lusigel'] = lg[:5]
    
    print('\n' + '=' * 60)
    print('SAMPLE dos primeiros 5 items por tabela:')
    print('=' * 60)
    for source, items in all_data.items():
        print(f'\n--- {source} ---')
        for item in items:
            print(f'  {item}')
    
    # Total
    total = sum(len(x) for x in [af, gg, mk, lg])
    print(f'\n\nTotal de items a inserir: {total}')
    
    # Salvar JSON para o script SQL consumir
    all_full = {
        'alpha_food': af,
        'gergran': gg,
        'makro': mk,
        'lusigel': lg,
    }
    with open('/tmp/parsed_data.json', 'w', encoding='utf-8') as f:
        json.dump(all_full, f, ensure_ascii=False, indent=2)
    print('JSON salvo em /tmp/parsed_data.json')
